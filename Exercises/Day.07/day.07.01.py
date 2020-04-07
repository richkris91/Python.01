# week = {'day': ['mo', 'tu', 'we', 'fe', 'fr', 'sa', 'su']}
# weekend = {'day': ['sa', 'su']}
# or element in week['day'[:]]:
#     if element in week['day'[:]] == element in weekend['day'[:]]:
#         print('OH YEAH')
#     else:
#         print('Oh no!')
import openpyxl
# excel = openpyxl.Workbook()
# excel.save("*file.xlsx")
file_name = "*file.xlsx"
excel = openpyxl.load_workbook(file_name)
print(excel.sheetnames)
sheet = excel.activerow = 1
column = 1
max_1 = int(input("By how much do you want to multiply: "))
for row in range(1, max_1 + 1):
    for column in range(1, max_1 + 1):
        cell = sheet.cell(row, column)
        cell.value = row * column
